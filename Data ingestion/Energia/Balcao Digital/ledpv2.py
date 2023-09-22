import openpyxl
import os
import sys
#print(os.listdir("."))
from sshtunnel import SSHTunnelForwarder
import mysql.connector
from mysql.connector import Error
#import colorama
import time

import configdb

mtn={"janeiro":"01", "fevereiro":"02", "março":"03", "abril":"04", "maio":"05", "junho":"06", "julho":"07",\
"agosto":"08", "setembro":"09", "outubro":"10", "novembro":"11", "dezembro":"12"}

try:
    print('Connecting to the MariaDB Database...')

    ssh_tunnel = SSHTunnelForwarder("195.23.9.30",
        ssh_username="ppimenta",
        ssh_private_key=r"C:\Users\Pedro\.ssh\id_rsa",
        remote_bind_address=("localhost", 3306)
    )

    ssh_tunnel.start()
    print(ssh_tunnel.local_bind_address) 

    connection = mysql.connector.connect(
        user=configdb.dbUserName,
        password=configdb.dbUPass,
        host="127.0.0.1",
        port = ssh_tunnel.local_bind_port,
        database="BAZE"
    )

    cursor = connection.cursor(dictionary = True)
    cursor = connection.cursor(buffered=True)

    """
    cursor.execute("SELECT * FROM baze21r order by id desc limit 1")
    result = cursor.fetchall()
    print("$",cursor.rowcount)
    print(result[0])
    print ("----- done")
    """
    print ("... connection successful.")
except Error as e:
    print('Connection Has Failed: ', e)
    sys.exit(0)


apaga=False

if apaga:
	sql = "delete  from Consumo15m where CPE = 'PT0002000033164293CK';"
	print (" -> ", sql)
	cursor.execute(sql)
	connection.commit()
	print(cursor.rowcount)


for i in os.listdir("."):
	#print (i, "existe.", end="")
	#f (left(i,5)="Consu"):
	if i[:5]=="Consu":
		print(i, " seleccionado.")
		start_time = time.time()

		# Open Workbook
		wb = openpyxl.load_workbook(filename=i, data_only=True)

		# Get All Sheets
		a_sheet_names = wb.sheetnames
		# Get first sheet
		o_sheet = wb[a_sheet_names[0]]
		#print(o_sheet)

		# Get Cell Values
		cpe = o_sheet['B3'].value
		ee="B6"
		if o_sheet.cell(row=10, column=1).value=="Data":
			ee="B7"
			
		
		mes =o_sheet[ee].value.split(" ")[0]
		#print (".%s." % (mes))
		ano = o_sheet[ee].value.split(" ")[1]
		mton = mtn[mes]
		sql = "delete from Consumo15m where cpe='"+cpe+"' and hora like '"+ano +"-"+ mton+"%'"
		print ("sql delete:", sql)
		#sys.exit(0)
		cursor.execute(sql)
		connection.commit()

		#sys.exit(0)
		print("CPE:%s, mês:%s" % (cpe, mes))
		

		persiste=True
		# Sheet Maximum filled Rows and columns
		print(o_sheet.max_row, " linhas por ", o_sheet.max_column , " colunas.")

		k=0
		s=0

		if persiste:
			# print(".%s." % (o_sheet.cell(row=10, column=1).value))
			linicio=10
			if o_sheet.cell(row=10, column=1).value=="Data":
				linicio=11

			k=0
			z=0
			for l in range(linicio,o_sheet.max_row+1):
				ac=o_sheet.cell(row=l, column=1)
				data =ac.value.replace("/", "-") + " " +o_sheet.cell(row=l, column=2).value+":00"
				#sql = "select id from Consumo15m where cpe='"+cpe+"' and hora='"+data+"'"
				#print ("sql: ", sql)
				#cursor.execute(sql)
				#connection.commit()
				#print("    -> existem", cursor.rowcount, " registos.")
				#if cursor.rowcount==0 or True:
				if  True:
					if linicio==10:
						consumo=o_sheet.cell(row=l, column=3).value
						sql = "insert into Consumo15m (cpe, hora, DadosdeConsumo) values ('"+cpe+"', '"+data+"', "+str(consumo)+")"
					else:
						#consumo=o_sheet.cell(row=l, column=3).value
						sql = "insert into Consumo15m (cpe, hora, PotActiva, PotReactIndut, PotReactCapac) Values\
						 ('"+cpe+"', '"+data+"', "+str(o_sheet.cell(row=l, column=3).value) +", "+\
						 str(o_sheet.cell(row=l, column=4).value) +", "+\
						 str(o_sheet.cell(row=l, column=5).value) +")"
				
					#print (" -> ", sql)
					#sys.exit(9)

					cursor.execute(sql)
					#connection.commit()
					k=k+1
					z=z+cursor.rowcount
					#print(cursor.rowcount, " leituras inseridas")
					if (k%53==0):
						print ("%0.4d " % (k), end="")
				else:
					if cursor.rowcount >1:
						result = cursor.fetchone()
						sql = "delete from Consumo15m where id="+str(result[0])
						#print("$",cursor.rowcount)
						#print(result[0])
						print ("sql: ", sql)
						cursor.execute(sql)
						connection.commit()
					print (" %4.4d" % (l), end="")

			print(connection.commit())
			print( "\n %d %d registos 15m adicionados em %.2f secs." %(k, z,  (time.time() - start_time)))
			print ("last sql:", sql)

print ("done.")
sys.exit(0)
